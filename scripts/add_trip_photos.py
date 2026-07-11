# -*- coding: utf-8 -*-
"""AGGIUNGE al sito le foto NUOVE messe in data/Viaggi/<AAAA - Zona - NomeBarca>/,
SENZA rigenerare le foto esistenti (preserva foto e tag gia' presenti).

Sicuro per costruzione:
- lavora sul site/data/trips.json AUTORITATIVO (non ri-deriva dalla sorgente Skipper
  Images, che e' incompleta e degraderebbe il sito);
- salta i placeholder numerici (01.jpg = copie dal sito) e i file gia' presenti
  (dedup aHash vs le foto del viaggio gia' online);
- un doppione con nome descrittivo ARRICCHISCE i tag della foto esistente (aggiunge,
  mai toglie);
- taggia le foto nuove leggendo i nomi dal file (persone + luogo);
- rigenera data/photo-tags.json DA trips.json (derivato) e poi aggiorna crew.json via
  build_crew.py (che preserva gli overlay 2026).

Uso:  python scripts/add_trip_photos.py
      python scripts/add_trip_photos.py --dry   (mostra cosa farebbe, non scrive)
"""
import openpyxl, collections, os, glob, re, unicodedata, json, sys, shutil, tempfile, subprocess
from PIL import Image, ImageOps

ROOT = r"c:\Users\Edo\OneDrive - Bologna Business School\1.Progettazione\99_AI_Workspace\35_SailingAgent"
DRY = "--dry" in sys.argv
CV = os.path.join(ROOT, "data", "Skipper CV Enriched.xlsx")
TRIPS = os.path.join(ROOT, "site", "data", "trips.json")
OUTIMG = os.path.join(ROOT, "site", "trips", "img")
VIAGGIDIR = os.path.join(ROOT, "data", "Viaggi")
PHOTOTAGS = os.path.join(ROOT, "data", "photo-tags.json")
IMG_EXT = (".jpg", ".jpeg", ".png", ".webp")
DEDUP_THR = 5

# --- registro passeggeri (per taggare i nomi nelle foto nuove) — come build_trips.py ---
_cvtmp = os.path.join(tempfile.gettempdir(), "nina_cv_add.xlsx")
try: shutil.copy(CV, _cvtmp)
except PermissionError:
    if not os.path.exists(_cvtmp): raise
wb = openpyxl.load_workbook(_cvtmp, data_only=True)
people = {}
for r in wb["Passengers"].iter_rows(min_row=2, values_only=True):
    if r[0] is None: continue
    people[int(r[0])] = {"cog": (r[1] or "").strip(), "nom": (r[2] or "").strip()}
part = collections.defaultdict(list)
for r in wb["Participations"].iter_rows(min_row=2, values_only=True):
    if r[0] is None or r[2] is None: continue
    try: t = int(r[0]); p = int(r[2])
    except: continue
    part[p].append(t)

def norm2(s): return "".join(c for c in unicodedata.normalize("NFD", (s or "").lower()) if c.isalnum())
def slug(s): return re.sub(r"[^a-z0-9]+", "-", unicodedata.normalize("NFKD", s.lower()).encode("ascii", "ignore").decode()).strip("-")
def norm(s): return "".join(c for c in unicodedata.normalize("NFD", s.lower()) if c.isalnum() or c == " ")
FIRSTFIX = {"jack": "giacomo", "bvernardo": "bernardo"}
ALIAS = {"edoardo-c": "edo-c", "gabriele-m": "gabri-m", "federico-b": "fede-b"}
MANUAL = {"edoardo casella": "edo-c"}
KW = {"grecia": "", "croazia": "", "spagna": "", "sardegna": "", "sicilia": "", "grenadine": "", "baleari": "", "corsica": "", "caraibi": "", "seychelles": "", "cicladi": "", "sporadi": "", "ionie": "", "dodecaneso": "", "egadi": ""}
STRIP = {"incidente", "incidenti", "naufragio", "di", "dei", "in", "porto", "trouble", "shooting", "pescatori", "napoletani", "agata", "leo", "e"}

def resolve_pid(fullname):
    toks = fullname.split()
    if len(toks) < 2: return None
    first = FIRSTFIX.get(norm2(toks[0]), norm2(toks[0])); last = norm2(toks[-1])
    if len(first) < 2 or len(last) < 2: return None
    cands = [pid for pid, pp in people.items()
             if norm2(pp["nom"]).startswith(first[:4]) and norm2(pp["cog"]).startswith(last[:4])]
    if not cands: return None
    return max(cands, key=lambda p: len(part.get(p, [])))
def crew_id_of(pid):
    pp = people[pid]; cid = slug(pp["nom"] + " " + pp["cog"][:1])
    return ALIAS.get(cid, cid)
def parse_people(basename):
    stem = os.path.splitext(basename)[0]; cids = []
    for seg in re.split(r"\s+-\s+", stem):
        s = re.sub(r"20[0-2][0-9](-\d{1,2})?", " ", seg).replace("-", " ")
        for k in KW: s = re.sub(r"(?i)\b" + k + r"\b", " ", s)
        s = re.sub(r"(?i)\b(i{1,4}|iv|v)\b", " ", s)
        s = re.sub(r"\([^)]*\)", " ", s); s = re.sub(r"\b\d+\b", " ", s)
        s = " ".join(w for w in s.split() if w.lower() not in STRIP).strip()
        if len(s.split()) < 2: continue
        nk = " ".join(norm(s).split())
        if nk in MANUAL: cids.append(MANUAL[nk]); continue
        pid = resolve_pid(s)
        if pid: cids.append(crew_id_of(pid))
    return list(dict.fromkeys(cids))

# --- immagini ---
def save_opt(im, path):
    mx = 1280
    if max(im.size) > mx:
        sc = mx / max(im.size); im = im.resize((round(im.width * sc), round(im.height * sc)), Image.LANCZOS)
    im.save(path, "JPEG", quality=82, optimize=True, progressive=True)
    return im
def ahash(im):
    px = im.convert("L").resize((8, 8), Image.LANCZOS).tobytes(); avg = sum(px) / 64
    v = 0
    for i, p in enumerate(px):
        if p >= avg: v |= 1 << i
    return v
def ham(a, b): return bin(a ^ b).count("1")

# --- trips.json autoritativo ---
doc = json.load(open(TRIPS, encoding="utf-8"))
trips_out = doc["trips"]
by_name = {slug(t["name"]): t for t in trips_out}

added, enriched, skipped_folders = [], [], []
for folder in sorted(glob.glob(os.path.join(VIAGGIDIR, "*"))):
    if not os.path.isdir(folder): continue
    files = [f for f in sorted(glob.glob(os.path.join(folder, "*")))
             if os.path.isfile(f) and os.path.splitext(f)[1].lower() in IMG_EXT
             and not re.fullmatch(r"\d+", os.path.splitext(os.path.basename(f))[0])]
    if not files: continue                     # solo placeholder numerici o vuota -> niente da aggiungere
    parts = os.path.basename(folder).split(" - ")
    t = by_name.get(slug(parts[-1].strip()))
    if t is None:
        skipped_folders.append((os.path.basename(folder), "cartella non mappata a un viaggio")); continue
    tid = t["id"]; zona = parts[1].strip() if len(parts) >= 3 else str(t.get("country", ""))
    d = os.path.join(OUTIMG, tid)
    # aHash delle foto gia' online del viaggio + prossimo numero libero
    existing = []
    nums = []
    for i, ph in enumerate(t["photos"]):
        m = re.search(r"/(\d+)\.jpg$", ph["src"]);  nums.append(int(m.group(1)) if m else 0)
        p = os.path.join(ROOT, "site", ph["src"].replace("/", os.sep))
        if os.path.isfile(p):
            try: existing.append((ahash(ImageOps.exif_transpose(Image.open(p)).convert("RGB")), i))
            except Exception: pass
    nextn = (max(nums) + 1) if nums else 1
    for f in files:
        b = os.path.basename(f)
        try: im = ImageOps.exif_transpose(Image.open(f)).convert("RGB")
        except Exception as e:
            skipped_folders.append((b, "immagine illeggibile: %s" % e)); continue
        h = ahash(im); cids = parse_people(b)
        di = next((idx for hh, idx in existing if ham(hh, h) <= DEDUP_THR), None)
        if di is not None:                     # doppione di una foto gia' online -> arricchisci i tag
            new = [c for c in cids if c not in t["photos"][di]["people"]]
            if new and not DRY: t["photos"][di]["people"].extend(new)
            if new: enriched.append((tid, b, new))
            continue
        fn = "%02d.jpg" % nextn; nextn += 1     # foto nuova -> ottimizza e appendi
        if not DRY:
            os.makedirs(d, exist_ok=True); im2 = save_opt(im, os.path.join(d, fn))
            w, hh = im2.width, im2.height
        else:
            w = hh = 0
        t["photos"].append({"src": "trips/img/%s/%s" % (tid, fn), "w": w, "h": hh,
                            "people": cids, "caption": ("%s %s" % (zona, t.get("year", ""))).strip()})
        existing.append((h, len(t["photos"]) - 1))
        added.append((tid, b, cids))
    t["n_photos"] = len(t["photos"])

# --- scrittura trips.json + rigenerazione derivati ---
if not DRY:
    json.dump(doc, open(TRIPS, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    by_person = collections.defaultdict(list)
    for t in trips_out:
        for p in t["photos"]:
            for cid in p["people"]:
                by_person[cid].append({"src": p["src"], "trip_id": t["id"], "caption": p["caption"]})
    json.dump({"generated_at": "2026-07-11", "by_person": by_person},
              open(PHOTOTAGS, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

# --- report ---
print(("DRY-RUN — " if DRY else "") + "foto aggiunte: %d, tag arricchiti: %d" % (len(added), len(enriched)))
for tid, b, c in added: print("  + %-12s %s  tag: %s" % (tid, b, ", ".join(c) or "-"))
for tid, b, c in enriched: print("  ~ %-12s %s  +tag: %s" % (tid, b, ", ".join(c)))
for b, why in skipped_folders: print("  ! %s -> %s" % (b, why))
if not DRY and (added or enriched):
    print("\nRigenero le gallery dei profili (build_crew.py)...")
    r = subprocess.run([sys.executable, os.path.join(ROOT, "scripts", "build_crew.py")],
                       capture_output=True, text=True)
    print(r.stdout.strip() or r.stderr.strip())
