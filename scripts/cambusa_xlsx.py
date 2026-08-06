"""Esporta il piano cambusa in un Excel condivisibile con l'equipaggio.

Fonte unica: build_plan() di cambusa_plan.py (stessi numeri della console).
PRIVACY (regola del progetto, il file gira su OneDrive condiviso):
  - niente nomi legati ad allergie: la riga "Pasta SENZA GLUTINE (nomi)"
    perde i nomi nel foglio; il dettaglio resta solo nella console locale
  - niente note dal questionario (preferenze/allergie): il foglio ha solo
    quantita' aggregate

Output: data/Cambusa 2026.xlsx (gitignorato: data/*.xlsx). Il file va poi
caricato su OneDrive; rigenerarlo quando cambiano equipaggio o questionari.

  python scripts/cambusa_xlsx.py
"""
from __future__ import annotations
import datetime as dt
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core import load, ROOT, aboard_on, nights_aboard, parse_date
from cambusa_plan import (ETICHETTE_VINO_BIANCO_LOCALI, build_plan,
                          detailed_lines, leg_boundaries, load_preferences,
                          name_to_crew_id)

OUT = ROOT / "data" / "Cambusa 2026.xlsx"

HEAD_FILL = PatternFill("solid", fgColor="1F4E5F")
HEAD_FONT = Font(bold=True, color="FFFFFF")
CAT_FILL = PatternFill("solid", fgColor="E8F0F2")

# categoria per voce: prefissi (il nome esatto puo' cambiare, es. quantita' tra
# parentesi); l'ordine della lista e' l'ordine dei blocchi nel foglio
CATEGORIE = [
    ("Liquidi", ("Acqua in bottiglia", "Coca Cola", "Estathé", "Succhi", "Sprite",
                 "Gatorade", "Birra analcolica", "Acqua tonica")),
    ("Alcolici", ("Birra (n)", "Aperol", "Prosecco", "Vino bianco", "Cannonau",
                  "Malvasia", "Mirto", "Filu")),
    ("Pasta e riso", ("Riso", "Pasta ")),
    ("Dispensa", ("Tonno", "Acciughe", "Legumi", "Pomodori", "Olio", "Sale grosso",
                  "Crackers", "Marmellata")),
    ("Colazione e snack", ("Caffe", "Caffe'", "Snack")),
    ("Freschi (mercato dell'8)", ("Pane (g)", "Verdura", "Frutta", "Carne", "Latte",
                                  "Formaggio", "Uova", "Ghiaccio")),
    ("Pulizia e consumabili", ("Carta igienica", "Rotoloni", "Detersivo", "Spugne",
                               "Panni", "Sacchi", "Sacchetti", "Sgrassatore",
                               "Detergente", "Sapone", "Gel igienizzante", "Pellicola",
                               "Tovaglioli", "Mollette", "Antizanzare", "Dopopuntura",
                               "Guanti", "Accendini")),
]


def categoria(voce: str) -> str:
    for cat, prefissi in CATEGORIE:
        if any(voce.startswith(p) for p in prefissi):
            return cat
    return "Altro"


def _san(voce: str) -> str:
    """La riga senza glutine perde i nomi (allergie mai fuori dal locale)."""
    return re.sub(r"^Pasta SENZA GLUTINE \(.*\)$", "Pasta SENZA GLUTINE", voce)


def _sheet_header(ws, cols, widths):
    for i, (name, w) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def build_xlsx() -> None:
    v = load()
    r = build_plan(v)
    g = r["spesa_grossa"]

    tutte_le_voci = {**g["non_deperibile_intera_crociera"], **g["deperibile_prima_tratta"]}
    righe = (detailed_lines(tutte_le_voci)
             + [{**x, "canale": "online"} for x in g["dispensa_essenziali"]]
             + [{**x, "canale": "online"} for x in g["pasta_dettaglio"]]
             + g["bevande_dettaglio"]
             + g["pulizia_consumabili"])
    # chi ha chiesto cosa (dal questionario): SOLO gusti — bevande e cibo
    # immancabile. Niente allergie con nomi: quelle le gestisce lo skipper.
    prefs = load_preferences()
    n2id = name_to_crew_id()
    chiesto_da: dict[str, list[str]] = {}
    immancabili: list[tuple[str, str]] = []
    for m in v["crew"]:
        p = prefs.get(n2id.get(m["name"], ""))
        if not p:
            continue
        for b in p["bibite"]:
            chiesto_da.setdefault(b, []).append(m["name"])
        if p["cibo_immancabile"]:
            immancabili.append((m["name"], p["cibo_immancabile"]))
    # Vermentino/Rose' corso = stessa quota "Vino bianco", etichetta preferita
    for et in ETICHETTE_VINO_BIANCO_LOCALI:
        for nome in chiesto_da.pop(et, []):
            chiesto_da.setdefault("Vino bianco", []).append(f"{nome} (etich. locale)")
    if "Birra" in chiesto_da:                    # label del form vs voce RATES
        chiesto_da.setdefault("Birra (n)", []).extend(chiesto_da.pop("Birra"))

    for x in righe:
        x["voce"] = _san(x["voce"])
        x["categoria"] = categoria(x["voce"])
        nomi = chiesto_da.get(x["voce"])
        if nomi:
            x["nota"] = f"per: {', '.join(sorted(nomi))}"

    wb = Workbook()

    # --- foglio 1: spesa grossa, ordinata per categoria (blocchi) poi canale
    ws = wb.active
    ws.title = "Spesa grossa 8-8"
    cols = ["Fatto", "Categoria", "Articolo", "Confezioni", "Formato",
            "Canale", "Q.ta' totale", "Note"]
    _sheet_header(ws, cols, [7, 22, 44, 11, 32, 11, 12, 28])
    ordine_cat = [c for c, _ in CATEGORIE] + ["Altro"]
    righe.sort(key=lambda x: (ordine_cat.index(x["categoria"]),
                              x["canale"] != "online", x["voce"].lower()))
    row, last_cat = 2, None
    for x in righe:
        if x["categoria"] != last_cat:
            last_cat = x["categoria"]
            c = ws.cell(row=row, column=2, value=last_cat.upper())
            c.font = Font(bold=True)
            for i in range(1, len(cols) + 1):
                ws.cell(row=row, column=i).fill = CAT_FILL
            row += 1
        ws.cell(row=row, column=2, value=x["categoria"])
        ws.cell(row=row, column=3, value=x["voce"])
        ws.cell(row=row, column=4, value=x["confezioni"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=x["formato"])
        ws.cell(row=row, column=6, value=x["canale"])
        q = x["quantita_totale"]
        ws.cell(row=row, column=7, value=q if q != "-" else None)
        if x.get("nota"):
            ws.cell(row=row, column=8, value=x["nota"]).alignment = Alignment(wrap_text=True)
        row += 1

    # --- foglio 2: chi c'e' a bordo per tratta — spiega PER CHI si compra e
    # perche' le quantita' sono quelle (scalano sulle persona-notti). Nomi brevi
    # pubblici e date imbarco/sbarco: gia' pubblicati sul sito (pagina arrivi).
    wsb = wb.create_sheet("Chi a bordo")
    legs = leg_boundaries(v)
    tratte = list(zip(legs[:-1], legs[1:]))
    fmt = lambda s: f"{parse_date(s).day}/{parse_date(s).month}"
    cols = ["Chi"] + [f"{fmt(d0)} → {fmt(d1)}" for d0, d1 in tratte] + ["Notti tot"]
    _sheet_header(wsb, cols, [20] + [12] * len(tratte) + [10])
    center = Alignment(horizontal="center")
    # righe di sintesi in alto: sono il moltiplicatore delle quantita'
    sintesi = [
        ("Persone a bordo", [len(aboard_on(v, d0)) for d0, _ in tratte], len(v["crew"])),
        ("Notti della tratta", [(parse_date(d1) - parse_date(d0)).days for d0, d1 in tratte],
         (parse_date(legs[-1]) - parse_date(legs[0])).days),
        ("Persona-notti", [len(aboard_on(v, d0)) * (parse_date(d1) - parse_date(d0)).days
                           for d0, d1 in tratte], sum(nights_aboard(m) for m in v["crew"])),
    ]
    row = 2
    for label, vals, tot in sintesi:
        wsb.cell(row=row, column=1, value=label).font = Font(bold=True)
        for j, x in enumerate(vals, start=2):
            wsb.cell(row=row, column=j, value=x).alignment = center
        wsb.cell(row=row, column=len(cols), value=tot).alignment = center
        for i in range(1, len(cols) + 1):
            wsb.cell(row=row, column=i).fill = CAT_FILL
        row += 1
    # matrice presenze: un pallino per ogni tratta in cui la persona e' a bordo
    for m in sorted(v["crew"], key=lambda m: (m["board"], m["leave"], m["name"])):
        wsb.cell(row=row, column=1, value=m["name"])
        for j, (d0, _) in enumerate(tratte, start=2):
            if parse_date(m["board"]) <= parse_date(d0) < parse_date(m["leave"]):
                wsb.cell(row=row, column=j, value="●").alignment = center
        wsb.cell(row=row, column=len(cols), value=nights_aboard(m)).alignment = center
        row += 1
    # cambi equipaggio: chi sale e chi scende a ogni data di confine
    row += 1
    wsb.cell(row=row, column=1, value="CAMBI EQUIPAGGIO").font = Font(bold=True)
    row += 1
    for d in legs:
        sale = sorted(m["name"] for m in v["crew"] if m["board"] == d)
        scende = sorted(m["name"] for m in v["crew"] if m["leave"] == d)
        parts = ([f"sale: {', '.join(sale)}"] if sale else []) + \
                ([f"scende: {', '.join(scende)}"] if scende else [])
        if not parts:
            continue
        wsb.cell(row=row, column=1, value=fmt(d)).font = Font(bold=True)
        c = wsb.cell(row=row, column=2, value=" · ".join(parts))
        wsb.merge_cells(start_row=row, start_column=2,
                        end_row=row, end_column=len(cols))
        c.alignment = Alignment(wrap_text=True, vertical="top")
        wsb.row_dimensions[row].height = 15 * (1 + (len(str(c.value)) // 80))
        row += 1

    # --- foglio 3: chi ha chiesto cosa (solo gusti dal questionario)
    wsr = wb.create_sheet("Richieste")
    _sheet_header(wsr, ["Chi", "Cibo immancabile"], [20, 90])
    row = 2
    for nome, cosa in sorted(immancabili):
        wsr.cell(row=row, column=1, value=nome)
        wsr.cell(row=row, column=2, value=cosa).alignment = Alignment(wrap_text=True)
        row += 1
    row += 1
    c = wsr.cell(row=row, column=1, value="Le bevande richieste da ciascuno sono nella colonna Note della spesa grossa.")
    c.font = Font(italic=True)
    row += 1
    wsr.cell(row=row, column=1,
             value="Solo gusti dal questionario — allergie e intolleranze le gestisce lo skipper a parte.").font = Font(italic=True)

    # --- foglio 4: rifornimenti ai cambi equipaggio (deperibili per tratta)
    ws2 = wb.create_sheet("Rifornimenti")
    _sheet_header(ws2, ["Fatto", "Tratta", "Persona-giorno", "Articolo", "Quantita'"],
                  [7, 26, 15, 30, 12])
    row = 2
    for leg in r["rifornimenti_successivi"]:
        for voce, q in leg["deperibile"].items():
            ws2.cell(row=row, column=2, value=leg["tratta"])
            ws2.cell(row=row, column=3, value=leg["persona_giorno"])
            ws2.cell(row=row, column=4, value=voce)
            ws2.cell(row=row, column=5, value=q)
            row += 1

    # --- foglio 3: leggimi
    ws3 = wb.create_sheet("Leggimi")
    ws3.column_dimensions["A"].width = 110
    note = [
        "CAMBUSA NINA 2026 — baseline di spesa condivisa",
        "",
        f"Spesa grossa il {g['giorno']} (primo imbarco): tutto il non deperibile per l'intera crociera",
        "+ i freschi della prima tratta. I fogli 'Rifornimenti' coprono i deperibili ai cambi equipaggio.",
        f"Budget stimato spesa grossa: ~{g['budget_stimato_eur']} EUR.",
        "",
        "Canale 'online' = ordinabile in anticipo (es. Conad Spesa Online, area Olbia-Tempio,",
        "consegna/ritiro prima dell'8). Canale 'di persona' = mercato/pescheria/enoteca l'8 mattina",
        "(freschi e specialita' locali).",
        "",
        "Il foglio 'Chi a bordo' spiega per chi si compra: persone e persona-notti per tratta",
        "(sono il moltiplicatore delle quantita') e i cambi equipaggio a ogni data.",
        "",
        "Come si usa: spunta la colonna 'Fatto' man mano; se cambi quantita' o aggiungi righe,",
        "scrivilo in Note cosi' lo skipper riallinea il piano.",
        "Le quantita' vengono dal questionario equipaggio (aggregate, pesate sulle notti a bordo",
        "di ciascuno, con margine di sicurezza). Niente nomi ne' dati personali in questo file:",
        "allergie e preferenze singole le gestisce lo skipper a parte.",
        "",
        "NB carta igienica: SOLO quella dissolvibile per WC marini (quella normale intasa le pompe).",
        "",
        f"Generato il {dt.date.today().isoformat()} da scripts/cambusa_xlsx.py — rigenerare se cambia l'equipaggio.",
    ]
    for i, t in enumerate(note, start=1):
        c = ws3.cell(row=i, column=1, value=t)
        if i == 1:
            c.font = Font(bold=True, size=13)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Scritto {OUT} — {len(righe)} righe spesa grossa, "
          f"{sum(len(l['deperibile']) for l in r['rifornimenti_successivi'])} righe rifornimenti")


if __name__ == "__main__":
    build_xlsx()
