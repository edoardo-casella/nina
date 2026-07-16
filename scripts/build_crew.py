# -*- coding: utf-8 -*-
# Dataset unico crew.json: 2026 crew (nickname) + alumni (registro), con giorni,
# viaggi, anni e grado navale (per seniority = giorni in barca).
import openpyxl, collections, json, re, unicodedata, io, os
# ROOT = radice del repo (genitore di scripts/), relativo alla posizione dello
# script: cosi' funziona anche dai git worktree fuori OneDrive, non solo qui.
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CV = os.path.join(ROOT, "data", "Skipper CV Enriched.xlsx")
import shutil, tempfile
_cvtmp = os.path.join(tempfile.gettempdir(), "nina_cv_read.xlsx")
try:
    shutil.copy(CV, _cvtmp)  # copia: il file OneDrive/Excel puo' essere lockato
except PermissionError:
    if not os.path.exists(_cvtmp): raise   # nessuna copia utilizzabile -> chiudi Excel
    print("CV lockato da Excel/OneDrive: uso la copia temp esistente", _cvtmp)
CREWJSON = os.path.join(ROOT, "site", "data", "crew.json")
# la zona "Inner Islands" (crociera Seychelles) va mostrata come "Seychelles"
ZONE_OVERRIDE = {"Inner Islands": "Seychelles"}
# Sorbe 2017: la rotta caraibica tocca Martinica (FR), Saint Vincent e Saint Lucia
COUNTRY_OVERRIDE = {"Saint Vincent - France": "Saint Vincent - France - Saint Lucia"}

# Soprannomi per profilo (crew_id -> lista di nick). Embedded qui (non in data/)
# perche' data/ e' condiviso via junction nei worktree: un file nuovo darebbe
# conflitti al pull. Attaccati come campo "nicks" a ogni persona in crew.json.
NICKS = {
    "andrea-a": ["Abbute Puche", "Immortale di Secondigliano", "Vulcan", "Idris", "Aibs", "Abba"],
    "bernardo-b": ["Berna"],
    "bianca-b": ["Bibi"],
    "edo-c": ["Edo"],
    "fede-b": ["Fede"],
    "fede-n": ["Fede"],
    "gabri-m": ["Gastone", "Gustavo", "Gustav"],
    "giacomo-n": ["Jack"],
    "ginevra-l": ["Gine", "Gi", "Ginx"],
    "giulia-n": ["Juliet", "Giuli"],
    "ilaria-m": ["Ila"],
    "lavinia-p": ["Peeiss"],
    "mati-m": ["Mati"],
    "riccardo-b": ["Branca"],
    "simona-a": ["Simo"],
    "alberto-b": ["Albi"],
    "alice-s": ["Spampi"],
    "angela-k": ["Angi"],
    "beatrice-f": ["Bea"],
    "belinda-b": ["Beli"],
    "carlo-a": ["Kampione"],
    "carlotta-d": ["Carlottina"],
    "cecilia-r": ["Ceci"],
    "claudio-c": ["Claude"],
    "enrico-g": ["Guercia"],
    "fabiana-f": ["Fabi"],
    "filippo-r": ["Filo"],
    "gabriele-b": ["Ga", "Il rosso"],
    "giacomo-b": ["Jack", "Biso"],
    "gianmarco-m": ["Mex"],
    "gualtiero-s": ["Saba", "Gualti", "Walzer"],
    "laura-c": ["Uein"],
    "leonardo-m": ["Masca", "Beppe", "Giuseppe", "Leo"],
    "luca-b": ["Giudeo"],
    "marco-t": ["Tomma", "Kung"],
    "margherita-c": ["Marghe"],
    "marta-a": ["Martolina"],
    "matilde-c": ["Mati"],
    "michele-m": ["Mike"],
    "nicola-r": ["Nick"],
    "paola-c": ["Pola"],
    "pietro-c": ["Callo"],
    "riccardo-t": ["Tugos", "Tugno", "Ricky"],
    "sara-b": ["Magu"],
    "silvano-b": ["il nano", "Nanu"],
    "tonino-l": ["Tony"],
    "umberto-f": ["Umbe", "Fasa"],
}

# Crew femminili (crew_id) -> per il "detto/detta" sui profili. Attaccato come "sex".
FEM = {
    "bianca-b", "fede-n", "ginevra-l", "giulia-n", "ilaria-m", "lavinia-p", "mati-m", "simona-a",
    "alice-s", "angela-k", "beatrice-f", "belinda-b", "carlotta-d", "cecilia-r", "chiara-f",
    "fabiana-f", "jennifer-s", "laura-c", "ludovica-r", "margherita-c", "marta-a", "martina-g",
    "martina-r", "matilde-c", "matilde-r", "monica-m", "paola-c", "rebecca-b", "rebecca-l",
    "sara-b", "simona-b", "sofia-m",
}

wb = openpyxl.load_workbook(_cvtmp, data_only=True)
people = {}
for r in wb["Passengers"].iter_rows(min_row=2, values_only=True):
    if r[0] is None: continue
    people[int(r[0])] = {"cog": (r[1] or "").strip(), "nom": (r[2] or "").strip()}
trip_days, trip_year, trip_nm = {}, {}, {}
trip_name, trip_zone, trip_country, trip_boat = {}, {}, {}, {}
for r in wb["Summary"].iter_rows(min_row=2, values_only=True):
    if r[0] in (None, "TOTAL"): continue
    try: t = int(float(r[0]))
    except: continue
    trip_year[t] = int(float(r[1])); trip_days[t] = float(r[10]); trip_nm[t] = int(float(r[9]))
    trip_name[t] = (r[8] or "").strip(); _z = (r[5] or "").strip(); trip_zone[t] = ZONE_OVERRIDE.get(_z, _z)
    _c = (r[4] or "").strip(); trip_country[t] = COUNTRY_OVERRIDE.get(_c, _c); trip_boat[t] = (r[6] or "").strip()
part = collections.defaultdict(list)       # pid -> [trip id]
roster = collections.defaultdict(list)     # trip id -> [pid]  (per la co-navigazione)
part_days, part_nm = {}, {}                 # (trip, pid) -> giorni/miglia PER-PERSONA
for r in wb["Participations"].iter_rows(min_row=2, values_only=True):
    if r[0] is None or r[2] is None: continue
    try: t = int(r[0]); p = int(r[2])
    except: continue
    part[p].append(t); roster[t].append(p)
    try: part_days[(t, p)] = float(r[6]) if r[6] is not None else None
    except: part_days[(t, p)] = None
    try: part_nm[(t, p)] = float(r[7]) if r[7] is not None else None
    except: part_nm[(t, p)] = None

def norm(s): return "".join(c for c in unicodedata.normalize("NFD", (s or "").lower()) if c.isalnum())
def slugify(s): return re.sub(r"[^a-z0-9]+", "-", unicodedata.normalize("NFKD", s.lower()).encode("ascii", "ignore").decode()).strip("-")

# I viaggi con avvicendamenti sono descritti in un sidecar testuale e versionabile.
# Il foglio Excel conserva il registro storico, ma non rappresenta quali persone
# abbiano realmente condiviso ciascun turno: senza questo overlay tutti i nomi del
# viaggio verrebbero considerati compagni tra loro.
GROUPS_FILE = os.path.join(ROOT, "data", "trip-crew-groups.json")
group_cfg = json.loads(io.open(GROUPS_FILE, encoding="utf-8").read()).get("trips", {}) if os.path.exists(GROUPS_FILE) else {}
crew_groups = {}  # trip numerico -> [{days, pids}], usato anche per la co-navigazione
cid_alias = {"edoardo-c": "edo-c", "gabriele-m": "gabri-m", "federico-b": "fede-b"}
pid_by_cid = {}
for pid, pp in people.items():
    cid = slugify(pp["nom"] + " " + pp["cog"][:1])
    pid_by_cid[cid_alias.get(cid, cid)] = pid

for trip_slug, spec in group_cfg.items():
    matches = [t for t, name in trip_name.items() if slugify(name) == trip_slug]
    if len(matches) != 1:
        raise ValueError(f"Configurazione equipaggi {trip_slug}: viaggio non trovato o ambiguo")
    t = matches[0]
    if spec.get("trip_id") is not None and int(spec["trip_id"]) != t:
        raise ValueError(f"Configurazione equipaggi {trip_slug}: trip_id {spec['trip_id']} non corrisponde a {t}")
    parsed, person_days = [], collections.Counter()
    for g in spec.get("groups", []):
        days = float(g["days"])
        if days <= 0:
            raise ValueError(f"{trip_slug}/{g.get('id')}: days deve essere positivo")
        pids = []
        for cid in g.get("crew", []):
            if cid == "edo-c":
                continue  # lo skipper e' fuori dal foglio Passengers ed e' gestito a parte
            pid = pid_by_cid.get(cid)
            if pid is None:
                raise ValueError(f"{trip_slug}/{g.get('id')}: crew id sconosciuto {cid}")
            pids.append(pid); person_days[pid] += days
        parsed.append({"days": days, "pids": pids})
    if abs(sum(g["days"] for g in parsed) - trip_days[t]) > 0.01:
        raise ValueError(f"{trip_slug}: la durata dei turni non coincide con i {trip_days[t]:g} giorni del viaggio")

    # La configurazione e' autoritativa: rimuove anche partecipazioni spurie dal
    # registro (per Alba 2020, Giacomo B.) e aggiunge eventuali persone mancanti.
    desired = set(person_days)
    for pid in list(roster.get(t, [])):
        if pid not in desired:
            part[pid] = [x for x in part.get(pid, []) if x != t]
            part_days.pop((t, pid), None); part_nm.pop((t, pid), None)
    roster[t] = list(dict.fromkeys(pid for g in parsed for pid in g["pids"]))
    for pid, days in person_days.items():
        if t not in part[pid]: part[pid].append(t)
        excel_days = part_days.get((t, pid))
        if excel_days is not None and excel_days < days:
            # Il registro Excel e' piu' preciso della somma dei turni (Alba 2020,
            # Alice S.: nel nucleo di entrambi i turni ma a bordo 15 giorni su 20).
            if part_nm.get((t, pid)) is None:
                part_nm[(t, pid)] = trip_nm[t] * excel_days / trip_days[t]
            continue
        part_days[(t, pid)] = days
        part_nm[(t, pid)] = trip_nm[t] * days / trip_days[t]
    crew_groups[t] = parsed

def pname(p):
    pp = people[p]; return (pp["nom"] + " " + (pp["cog"][:1] + "." if pp["cog"] else "")).strip() or pp["cog"] or ("#" + str(p))

def pdays(t, p): return part_days.get((t, p)) if part_days.get((t, p)) is not None else trip_days.get(t, 0)
def pnm(t, p):   return part_nm.get((t, p))   if part_nm.get((t, p))   is not None else trip_nm.get(t, 0)

def stats(p):  # giorni/miglia PER-PERSONA (alcuni sono a bordo meno del totale viaggio)
    ts = part.get(p, [])
    days = round(sum(pdays(t, p) for t in ts))
    nm = round(sum(pnm(t, p) for t in ts))
    ys = [trip_year.get(t) for t in ts if trip_year.get(t)]
    return {"trips": len(ts), "days": days, "nm": nm, "first": min(ys) if ys else None, "last": max(ys) if ys else None}

def trips_list_for(ts, pid=None):  # log viaggi cliccabile: id-slug (parita' con build_trips), + boat/nm/days (per pace/icona) + pdays (giorni a bordo di QUESTA persona; None -> intero viaggio)
    rows = []
    for t in sorted(ts, key=lambda t: (trip_year.get(t, 0), t)):
        row = {"id": slugify(trip_name.get(t, "")) or ("t" + str(t)),
               "year": trip_year.get(t), "zone": trip_zone.get(t, ""), "country": trip_country.get(t, ""),
               "boat": trip_boat.get(t, ""), "nm": trip_nm.get(t, 0), "days": round(trip_days.get(t, 0)),
               "pdays": round(pdays(t, pid)) if pid is not None else round(trip_days.get(t, 0))}
        # Per chi ha partecipato solo a un turno, conserva anche le miglia
        # personali. Per i viaggi interi `nm` e' gia' il valore corretto.
        if pid is not None and row["pdays"] < row["days"]:
            row["pnm"] = round(pnm(t, pid))
        rows.append(row)
    return rows

def companions_for(p):  # con chi ha navigato di piu' (peso: n. viaggi condivisi, tie-break giorni)
    if p is None: return []
    cnt, dsum, nsum = collections.Counter(), collections.Counter(), collections.Counter()
    for t in part.get(p, []):
        if t in crew_groups:
            # Conta solo le persone presenti nello stesso turno. Chi era nel core
            # di entrambi i turni condivide 20 giorni; gli altri soltanto i 10
            # giorni del proprio equipaggio. Il viaggio condiviso conta una volta.
            # Le miglia di ogni turno sono proporzionali alla sua durata.
            shared_days, shared_nm = collections.Counter(), collections.Counter()
            for g in crew_groups[t]:
                if p in g["pids"]:
                    g_nm = trip_nm.get(t, 0) * g["days"] / trip_days[t] if trip_days.get(t) else 0
                    for q in g["pids"]:
                        if q != p:
                            shared_days[q] += g["days"]; shared_nm[q] += g_nm
            for q, days in shared_days.items():
                # cap alla partecipazione reale: chi ha fatto meno giorni dei suoi
                # turni (registro Excel) non puo' condividere piu' di quelli
                cnt[q] += 1
                dsum[q] += round(min(days, pdays(t, p), pdays(t, q)))
                nsum[q] += round(min(shared_nm[q], pnm(t, p), pnm(t, q)))
        else:
            for q in roster.get(t, []):
                if q == p: continue
                cnt[q] += 1
                dsum[q] += round(min(pdays(t, p), pdays(t, q)))
                nsum[q] += round(min(pnm(t, p), pnm(t, q)))
    ranked = sorted(cnt, key=lambda q: (cnt[q], dsum[q]), reverse=True)  # TUTTI i co-naviganti (filtro id dopo l'ordinamento, non prima: altrimenti gli stub tagliavano il conteggio)
    res = []
    # lo skipper e' a bordo di OGNI crociera -> compagno di navigazione n.1 di chiunque
    # (fa il viaggio intero: le miglia condivise coincidono con le miglia personali di p)
    ts = part.get(p, [])
    if ts:
        res.append({"id": "edo-c", "name": id2name.get("edo-c", "Edo C"),
                    "trips": len(ts), "days": round(sum(pdays(t, p) for t in ts)),
                    "nm": round(sum(pnm(t, p) for t in ts))})
    for q in ranked:
        qid = pid2id.get(q)
        if qid: res.append({"id": qid, "name": id2name.get(qid, ""), "trips": cnt[q], "days": dsum[q], "nm": nsum[q]})
    return res

def match(name):  # voyage nickname -> registry pid
    parts = name.strip().split()
    if len(parts) < 2: return None
    first = norm(parts[0]); si = norm(parts[-1])[:1]
    cands = [pid for pid, pp in people.items() if norm(pp["nom"]).startswith(first[:4]) and norm(pp["cog"])[:1] == si]
    return max(cands, key=lambda p: len(part.get(p, []))) if cands else None

# --- DUE CARRIERE DI GRADO --------------------------------------------------
# COMANDANTI: chi ha preso il comando di una barca -> carriera "capitano" (oro).
#   Anche al primo comando e' gia' comandante (Capitano di corvetta).
# ABILITATI: ha la patente ma non ha ancora comandato -> grado "ponte".
# MARINAI: tutti gli altri -> carriera d'equipaggio (argento), tetto Tenente di vascello.
COMMANDERS = {"edo-c", "claudio-c", "marco-t", "michele-m", "bernardo-b"}
ABILITATI = {"riccardo-b", "ilaria-m", "gabri-m"}
# esperienza ESTERNA ai viaggi con la Niña (altri comandi, barca propria, ecc.):
# giorni "virtuali" aggiunti SOLO al calcolo del grado (le statistiche reali
# di giorni/miglia restano quelle effettive). Es. Michele = super-commodoro.
EXT_DAYS = {"michele-m": 600, "bernardo-b": 55}
GRADES = {
    "marinaio": [
        (0, "mozzo", "Mozzo"), (7, "marinaio", "Marinaio"),
        (20, "marinaio-scelto", "Marinaio scelto"), (45, "nostromo", "Nostromo"),
        (80, "sottoten", "Sottotenente di vascello"), (130, "tenente", "Tenente di vascello"),
    ],
    "comandante": [
        (0, "cap-corvetta", "Capitano di corvetta"), (70, "cap-fregata", "Capitano di fregata"),
        (140, "cap-vascello", "Capitano di vascello"), (220, "contrammiraglio", "Contrammiraglio"),
        (280, "ammiraglio", "Ammiraglio"), (560, "grande-ammiraglio", "Grande Ammiraglio"),
    ],
    "abilitato": [(0, "abilitato", "Abilitato al comando")],
}
def track_of(cid, me=False):
    if me or cid in COMMANDERS: return "comandante"
    return "marinaio"   # gli ABILITATI restano marinai (tengono il grado) + flag abilitato
def grade_of(track, days):
    ladder = GRADES.get(track, GRADES["marinaio"])
    g = ladder[0]
    for step in ladder:
        if days >= step[0]: g = step
    return g   # (min_days, id, label)

cur = json.loads(io.open(CREWJSON, encoding="utf-8").read())
# ricostruisce la lista membri 2026 dal file unificato (people con crew2026),
# preservando gli overlay manuali (board/leave/role/photo/nick/note/link).
# NOTA: la scheda personale `q` NON esce più nel crew.json — vive nella tabella
# Supabase `profiles` (pubblica in lettura, editabile da profilo.html).
# board/leave sono tornate pubbliche per decisione di Edo (2026-07-16).
if "members" in cur:
    members = cur["members"]
else:
    members = [{"id": p["id"], "name": p["name"], **p.get("crew2026", {})}
               for p in cur["people"] if "crew2026" in p]
# board/leave: fonte autoritativa voyage.json (il crew.json corrente potrebbe
# non averle: sono state riservate per un giorno e poi ri-pubblicate)
try:
    _voy = json.load(io.open(os.path.join(ROOT, "data", "voyage.json"), encoding="utf-8"))
    _bl = {c["name"]: (c["board"], c["leave"]) for c in _voy.get("crew", [])
           if c.get("board") and c.get("leave")}
except Exception:
    _bl = {}
for _m in members:
    if _m["name"] in _bl:
        _m["board"], _m["leave"] = _bl[_m["name"]]
used_pids = set()
id2pid = {}   # crew_id -> registry pid (per companions)
out = []
TOTAL_DAYS = round(sum(trip_days.values())); TOTAL_NM = sum(trip_nm.values())  # Edoardo
for m in members:
    nm = m["name"]
    if nm.split()[0].lower() in ("edo", "edoardo"):
        st = {"trips": 25, "days": TOTAL_DAYS, "nm": TOTAL_NM, "first": 2011, "last": 2024}; me = True
        tl = trips_list_for(list(trip_year.keys()))   # lo skipper c'e' su tutti i viaggi
    else:
        pid = match(nm); me = False
        if pid: used_pids.add(pid); id2pid[m["id"]] = pid; st = stats(pid); tl = trips_list_for(part.get(pid, []), pid)
        else: st = {"trips": 0, "days": 0, "nm": 0, "first": None, "last": None}; tl = []
    crew2026 = {k: m[k] for k in ("board", "leave", "role", "photo", "nick", "note", "link") if k in m}
    tr = track_of(m["id"], me); ext = EXT_DAYS.get(m["id"], 0)
    _, rid, rlab = grade_of(tr, st["days"] + ext)
    out.append({"id": m["id"], "name": nm, **st, "rank": rid, "rank_label": rlab, "track": tr,
                **({"ext_days": ext} if ext else {}),
                **({"abilitato": True} if m["id"] in ABILITATI else {}),
                **({"me": True} if me else {}), "trips_list": tl, "crew2026": crew2026})
# alumni: registro non nel crew 2026
for p in part:
    if p in used_pids: continue
    pp = people[p]
    if len((pp.get("nom") or "")) < 2 or not pp.get("cog"): continue  # scarta voci-stub
    st = stats(p); nm = pname(p)
    pid_id = slugify(nm.replace(".", "")) or ("p" + str(p))
    id2pid[pid_id] = p
    tr = track_of(pid_id); ext = EXT_DAYS.get(pid_id, 0)
    _, rid, rlab = grade_of(tr, st["days"] + ext)
    out.append({"id": pid_id, "name": nm, **st, "rank": rid, "rank_label": rlab, "track": tr,
                **({"ext_days": ext} if ext else {}),
                **({"abilitato": True} if pid_id in ABILITATI else {}),
                "trips_list": trips_list_for(part.get(p, []), p)})

# cani di bordo: non sono nel registro Passengers, li aggiungiamo a mano come profili
# (pet=True li tiene fuori dalla classifica ciurma; avatar auto-agganciato sotto da
# site/crew/img/<id>.jpg). Presenti sull'Ondine, Croazia 2018.
_DOG_TRIP = {"id": "ondine", "year": 2018, "zone": "Croazia CZ", "country": "Croazia",
             "boat": "Sun Odyssey 519", "nm": 385, "days": 11, "pdays": 11}
DOGS = [
    {"id": "agata", "name": "Agata", "pet": True, "track": "marinaio", "rank": "mozzo", "rank_label": "Mozzo",
     "trips": 1, "days": 11, "nm": 385, "first": 2018, "last": 2018,
     "trips_list": [dict(_DOG_TRIP)],
     "bio": {"epithet": "Mascotte di bordo", "html":
             "<p>Una delle due mascotte a quattro zampe di casa. La Croazia 2018 è stata il suo "
             "battesimo del mare: mare calmo, ritmi lenti e le prime lezioni di rollio.</p>"}},
    {"id": "leo", "name": "Leo", "pet": True, "track": "marinaio", "rank": "mozzo", "rank_label": "Mozzo",
     "trips": 1, "days": 11, "nm": 385, "first": 2018, "last": 2018,
     "trips_list": [dict(_DOG_TRIP)],
     "bio": {"epithet": "Mascotte di bordo", "html":
             "<p>L'altra metà del duo di bordo. Croazia 2018, prima crociera in famiglia: acque "
             "tranquille per imparare a stare in coperta senza perdere l'equilibrio.</p>"}},
]
out.extend(DOGS)

# foto taggate per persona (sidecar generato da build_trips.py) -> campo photos
PHOTOTAGS = os.path.join(ROOT, "data", "photo-tags.json")
by_person = {}
if os.path.exists(PHOTOTAGS):
    by_person = json.loads(io.open(PHOTOTAGS, encoding="utf-8").read()).get("by_person", {})
else:
    print("ATTENZIONE: data/photo-tags.json assente -> nessuna foto sui profili. Esegui prima build_trips.py.")
# ritratti scritti a mano (data/crew-bios.json) -> bio; tag di profilo (data/crew-tags.json)
BIOS = os.path.join(ROOT, "data", "crew-bios.json")
bios = json.loads(io.open(BIOS, encoding="utf-8").read()).get("bios", {}) if os.path.exists(BIOS) else {}
TAGS = os.path.join(ROOT, "data", "crew-tags.json")
tagmap = json.loads(io.open(TAGS, encoding="utf-8").read()).get("tags", {}) if os.path.exists(TAGS) else {}
pid2id = {v: k for k, v in id2pid.items()}          # per companions_for
id2name = {x["id"]: x["name"] for x in out}
n_photos = n_bios = n_tags = n_comp = n_avatar = n_nicks = 0
for x in out:
    ph = by_person.get(x["id"], [])
    if ph:
        x["photos"] = ph; n_photos += len(ph)
    b = bios.get(x["id"])
    if b:
        x["bio"] = b; n_bios += 1
    tg = tagmap.get(x["id"])
    if tg:
        x["tags"] = tg; n_tags += 1
    nk = NICKS.get(x["id"])
    if nk:
        x["nicks"] = nk; n_nicks += 1
    x["sex"] = "f" if x["id"] in FEM else "m"
    comp = companions_for(id2pid.get(x["id"]))
    if comp:
        x["companions"] = comp; n_comp += 1
    # avatar: ritratto in site/crew/img/<id>.jpg (auto-agganciato per gli alumni,
    # cioe' i membri senza foto nell'overlay crew2026)
    av = os.path.join(ROOT, "site", "crew", "img", x["id"] + ".jpg")
    if os.path.exists(av) and not x.get("crew2026", {}).get("photo"):
        x["photo"] = "crew/img/" + x["id"] + ".jpg"; n_avatar += 1

out.sort(key=lambda x: (-x["days"], -x["trips"]))
data = {"generated_at": "2026-07-11",
        "ranks": {rid: lab for ladder in GRADES.values() for _, rid, lab in ladder},
        "n_total": len(out),
        "people": out}
io.open(CREWJSON, "w", encoding="utf-8").write(json.dumps(data, ensure_ascii=False, indent=1))
print("crew.json:", len(out), "persone. TOTAL_DAYS(Edo)=", TOTAL_DAYS,
      "| foto:", n_photos, "su", sum(1 for x in out if x.get("photos")), "pers.",
      "| bio:", n_bios, "| tags:", n_tags, "| nicks:", n_nicks, "| companions:", n_comp,
      "| trips_list:", sum(1 for x in out if x.get("trips_list")))
rc = collections.Counter(x["rank"] for x in out)
print("distribuzione gradi:", dict(rc))
for x in out[:20]:
    tag = " [2026]" if "crew2026" in x else ""
    print(f"  {x['days']:3d}gg {x['trips']:2d}tr  {x['rank']:8s}  {x['name']}{tag}")
