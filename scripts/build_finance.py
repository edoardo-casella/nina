"""Riepilogo finanziario PERSONALE del viaggio → Supabase (tab "Il tuo viaggio").

Legge data/Summer 26.xlsx (gitignorato, SOLO locale) e carica su Supabase una
riga voyage_finance per membro: cabina, quota barca, extra, transazioni, saldo.
RLS lato Supabase: ogni membro legge SOLO la propria riga (o l'admin).

Il ponte nomi Excel → crew_id sta in data/finance-map.local.json (gitignorato:
contiene i nomi completi). Nei payload le controparti compaiono SEMPRE col nome
pubblico ("Bernardo B"), mai col nome completo — belt & braces sul repo pubblico.

Fonti dentro l'Excel (fa fede Bonifici per le quote, decisione 2026-07-10):
  Bonifici     quota dovuta/finale, versato, delta da versare
  Transazioni  sez. 3 saldi per soggetto + sez. 4 registro movimenti
  Extra        catalogo + riparto per persona + notti S1/S2/S3 (autorita' notti)
  Cabine       matrice per-notte → intervalli cabina + compagni + date imbarco/sbarco
  Quote        date settimane + tipo cabina per settimana

Uso:
  python scripts/build_finance.py --dry-run          # riassunto + controlli, niente push
  python scripts/build_finance.py --dry-run --json   # anche i payload completi
  python scripts/build_finance.py                    # push (env SUPABASE_URL + SUPABASE_SERVICE_KEY)

MAI in CI (l'Excel non esiste sul runner): run manuale dopo ogni modifica al file.
L'Excel aperto in Excel e' lockato → si legge sempre da una copia temporanea.
"""
from __future__ import annotations
import argparse, datetime as dt, json, os, shutil, sys, tempfile, urllib.request
from pathlib import Path

import openpyxl

import core

XLSX = core.DATA / "Summer 26.xlsx"
MAP = core.DATA / "finance-map.local.json"
CREW = core.ROOT / "site" / "data" / "crew.json"

# Cosa comprende ogni extra (testi generici, ok nel repo — le CIFRE restano su
# Supabase). Aggiornare qui quando si aggiunge un extra nel foglio.
EXTRA_NOTES = {
    "Starter Pack": "Pacchetto obbligatorio del charter: pulizie finali, gas e dotazioni di bordo.",
    "Sup": "Noleggio SUP per tutta la crociera.",
    "Cauzione": "Cauzione del charter.",
    "National Park": "Permessi del parco nazionale (arcipelago di La Maddalena).",
    "Lenzuola": "Cambio lenzuola e biancheria — ripartito sulle notti a bordo.",
    "Early Checkin": "Imbarco anticipato della settimana 1.",
}

ERRORS: list[str] = []
WARNINGS: list[str] = []


def err(msg: str) -> None: ERRORS.append(msg)
def warn(msg: str) -> None: WARNINGS.append(msg)


def norm(s) -> str:
    """Normalizza un nome/token: spazi, punto NON rimosso (distingue 'Ilaria C.'
    da 'Ilaria M.'), via il '⚠' della matrice Cabine e gli spazi doppi."""
    return " ".join(str(s or "").replace("⚠", " ").split())


def num(x) -> float:
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def r2(x) -> float:
    return round(num(x) + 0.0, 2)   # +0.0 evita il -0.0


def iso(d) -> str | None:
    return d.date().isoformat() if isinstance(d, dt.datetime) else (
        d.isoformat() if isinstance(d, dt.date) else None)


def open_wb():
    """L'xlsx aperto in Excel/OneDrive e' lockato (PermissionError): si copia
    in una temp dir e si legge la copia (= ultimo salvataggio)."""
    tmp = Path(tempfile.mkdtemp(prefix="nina-fin-")) / XLSX.name
    shutil.copy2(XLSX, tmp)
    return openpyxl.load_workbook(tmp, read_only=True, data_only=True)


def find_header(rows: list[tuple], *first_cells: str) -> int:
    """Indice (0-based) della riga il cui inizio combacia con first_cells."""
    want = [norm(c) for c in first_cells]
    for i, r in enumerate(rows):
        if [norm(c) for c in r[:len(want)]] == want:
            return i
    sys.exit(f"Header {first_cells} non trovato nel foglio.")


# ---------- mappa nomi ----------
def load_map() -> dict:
    if not MAP.exists():
        sys.exit(f"Manca {MAP} — e' il ponte nomi Excel → crew_id (vedi header dello script).")
    doc = json.loads(MAP.read_text(encoding="utf-8"))
    people = doc["participants"]
    crew_ids = {p["id"] for p in json.loads(CREW.read_text(encoding="utf-8"))["people"]
                if p.get("crew2026")}
    by_id, by_name, by_alias = {}, {}, {}
    for p in people:
        if p["crew_id"] and p["crew_id"] not in crew_ids:
            err(f"finance-map: crew_id '{p['crew_id']}' non esiste in crew.json (crew2026)")
        if p["crew_id"] is None and not p.get("skip_reason"):
            err(f"finance-map: '{p['excel_name']}' ha crew_id null senza skip_reason")
        by_id[p["excel_id"]] = p
        by_name[norm(p["excel_name"])] = p
        for a in [p["excel_name"], p["public_name"], *p.get("aliases", [])]:
            by_alias[norm(a)] = p
    return {"by_id": by_id, "by_name": by_name, "by_alias": by_alias}


# ---------- reader per foglio ----------
def read_bonifici(ws, m) -> dict[int, dict]:
    rows = list(ws.iter_rows(values_only=True))
    h = find_header(rows, "ID", "Partecipante", "Quota dovuta €")
    out, tot = {}, None
    for r in rows[h + 1:]:
        name = norm(r[1])
        if name == "TOTALE":
            tot = {"quota_finale": r2(r[3]), "delta": r2(r[6])}
            break
        if r[0] is None or not name:
            continue
        p = m["by_name"].get(name)
        if not p:
            err(f"Bonifici: partecipante '{name}' non in finance-map"); continue
        if int(num(r[0])) != p["excel_id"]:
            err(f"Bonifici: ID {r[0]} per '{name}' ≠ excel_id {p['excel_id']} in finance-map")
        out[p["excel_id"]] = {"quota_dovuta": r2(r[2]), "quota_finale": r2(r[3]),
                              "versato": r2(r[4]), "delta": r2(r[6])}
    if tot:
        for k in ("quota_finale", "delta"):
            s = round(sum(v[k] for v in out.values()), 2)
            if abs(s - tot[k]) > 0.01:
                err(f"Bonifici: somma {k} {s} ≠ TOTALE {tot[k]}")
    return out


def read_saldi(ws, m) -> dict[int, dict]:
    rows = list(ws.iter_rows(values_only=True))
    h = find_header(rows, "ID", "Soggetto", "Quota barca €")
    out = {}
    for r in rows[h + 1:]:
        name = norm(r[1])
        if name == "TOTALE":
            break
        if r[0] is None or not name or name == "0":
            continue
        p = m["by_name"].get(name)
        if not p:
            err(f"Transazioni/saldi: soggetto '{name}' non in finance-map"); continue
        out[p["excel_id"]] = {"quota_barca": r2(r[2]), "extra": r2(r[3]), "dovuto": r2(r[4]),
                              "crediti": r2(r[5]), "versato": r2(r[6]), "ricevuto": r2(r[7]),
                              "saldo": r2(r[8]), "stato": norm(r[9])}
    return out


def read_movimenti(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    h = find_header(rows, "N.", "Data", "Da (ID)")
    out = []
    for r in rows[h + 1:]:
        if r[0] is None:
            break
        out.append({"n": int(num(r[0])), "date": iso(r[1]),
                    "from_id": int(num(r[2])), "to_id": int(num(r[4])),
                    "kind": norm(r[6]), "amount": r2(r[7]),
                    "note": norm(r[8]) or None})
    return out


def read_extra(ws, m) -> tuple[list[dict], dict[int, dict]]:
    rows = list(ws.iter_rows(values_only=True))
    h = find_header(rows, "N.", "Data", "Descrizione")
    catalog = []                                   # posizione 0..11 = extra n.1..12
    for r in rows[h + 1:]:
        if r[0] is None or norm(r[2]) == "TOTALE EXTRA":
            break
        catalog.append({"name": norm(r[2]) or None, "total": r2(r[3]),
                        "criterion": norm(r[4]) or None})
    hm = find_header(rows, "ID", "Partecipante", "Notti tot.")
    per = {}
    for r in rows[hm + 1:]:
        name = norm(r[1])
        if name == "TOTALE":
            break
        if r[0] is None or not name:
            continue
        p = m["by_name"].get(name)
        if not p:
            err(f"Extra: partecipante '{name}' non in finance-map"); continue
        shares = [r2(r[6 + k]) for k in range(len(catalog))]
        per[p["excel_id"]] = {
            "nights": {"tot": int(num(r[2])), "s1": int(num(r[3])),
                       "s2": int(num(r[4])), "s3": int(num(r[5]))},
            "items": [{"name": c["name"], "criterion": c["criterion"], "total": c["total"],
                       "share": s, "note": EXTRA_NOTES.get(c["name"] or "")}
                      for c, s in zip(catalog, shares) if c["name"] and s > 0.005],
            "total": r2(r[6 + len(catalog)] if len(r) > 6 + len(catalog) else
                        sum(shares)),
        }
    return catalog, per


def read_cabine(ws, m) -> dict[int, list[dict]]:
    """Matrice per-notte → per persona: intervalli {from,to,nights,cabin,shared_with}.
    'to' e' l'ULTIMA NOTTE (lo sbarco e' la mattina dopo)."""
    rows = list(ws.iter_rows(values_only=True))
    h = find_header(rows, "N.", "Cabina")
    dates = [iso(c) for c in rows[h][2:] if c is not None]
    nights = {}                                    # excel_id -> {date: (cabin, frozenset ids)}
    for r in rows[h + 1:]:
        cabin = norm(r[1])
        if r[0] is None or not cabin:
            break
        for di, cell in enumerate(r[2:2 + len(dates)]):
            if not cell or not norm(cell):
                continue
            ids = []
            for tok in str(cell).split("+"):
                p = m["by_alias"].get(norm(tok))
                if not p:
                    err(f"Cabine: token '{tok.strip()}' (cabina {cabin}, {dates[di]}) non risolto dalla finance-map")
                    continue
                ids.append(p["excel_id"])
            for pid in ids:
                nights.setdefault(pid, {})[dates[di]] = (cabin, frozenset(ids))
    out = {}
    day1 = dt.timedelta(days=1)
    for pid, byday in nights.items():
        ivals = []
        for d in sorted(byday):
            cabin, ids = byday[d]
            mates = sorted(m["by_id"][i]["public_name"] for i in ids if i != pid)
            prev = ivals[-1] if ivals else None
            contiguous = prev and (dt.date.fromisoformat(prev["to"]) + day1
                                   == dt.date.fromisoformat(d))
            if contiguous and prev["cabin"] == cabin and prev["shared_with"] == mates:
                prev["to"] = d; prev["nights"] += 1
            else:
                ivals.append({"from": d, "to": d, "nights": 1,
                              "cabin": cabin, "shared_with": mates})
        out[pid] = ivals
    return out


def read_quote(ws, m) -> tuple[list[dict], dict[int, dict]]:
    rows = list(ws.iter_rows(values_only=True))
    labels = {}
    for r in rows:
        lab = norm(r[1])
        for key, pre in [("start", "Data inizio"), ("c1", "Data cambio 1"),
                         ("c2", "Data cambio 2"), ("end", "Data fine")]:
            if lab.startswith(pre):
                labels[key] = iso(r[2])
    weeks = [{"n": 1, "from": labels.get("start"), "to": labels.get("c1")},
             {"n": 2, "from": labels.get("c1"), "to": labels.get("c2")},
             {"n": 3, "from": labels.get("c2"), "to": labels.get("end")}]
    h = find_header(rows, "ID", "Partecipante", "Cabina S1 (08–15)")
    types = {}
    for r in rows[h + 1:]:
        name = norm(r[1])
        if r[0] is None or not name:
            continue
        p = m["by_name"].get(name)
        if not p:
            err(f"Quote: partecipante '{name}' non in finance-map"); continue
        types[p["excel_id"]] = {1: norm(r[2]) or None, 2: norm(r[3]) or None,
                                3: norm(r[4]) or None}
    return weeks, types


# ---------- assemblaggio ----------
def week_of(date: str, weeks: list[dict]) -> int:
    for w in weeks:
        if w["from"] and w["to"] and w["from"] <= date < w["to"]:
            return w["n"]
    return 0


def build_payloads(wb, m) -> tuple[dict[str, dict], list[str]]:
    bon = read_bonifici(wb["Bonifici"], m)
    sal = read_saldi(wb["Transazioni"], m)
    mov = read_movimenti(wb["Transazioni"])
    _, extra = read_extra(wb["Extra"], m)
    cab = read_cabine(wb["Cabine"], m)
    weeks, ctypes = read_quote(wb["Quote"], m)

    voyage = {c["name"]: c for c in core.load().get("crew", [])}
    now = dt.datetime.now().astimezone().isoformat(timespec="seconds")
    payloads, skipped = {}, []
    for p in m["by_id"].values():
        pid, cid = p["excel_id"], p["crew_id"]
        if cid is None:
            skipped.append(f"{p['public_name']} — {p.get('skip_reason', '?')}")
            continue
        ivals = cab.get(pid, [])
        ex = extra.get(pid) or {"nights": {"tot": 0, "s1": 0, "s2": 0, "s3": 0},
                                "items": [], "total": 0.0}
        n_cab = sum(i["nights"] for i in ivals)
        if n_cab != ex["nights"]["tot"]:
            err(f"{p['public_name']}: notti in Cabine ({n_cab}) ≠ Notti tot. in Extra ({ex['nights']['tot']})")
        board = ivals[0]["from"] if ivals else None
        leave = (dt.date.fromisoformat(ivals[-1]["to"]) + dt.timedelta(days=1)).isoformat() if ivals else None
        v = voyage.get(p["public_name"])
        if v and (v.get("board") != board or v.get("leave") != leave):
            warn(f"{p['public_name']}: date Excel {board}→{leave} ≠ voyage.json "
                 f"{v.get('board')}→{v.get('leave')} (fa fede l'Excel; aggiornare voyage.json)")
        for i in ivals:
            wk = week_of(i["from"], weeks)
            i["type"] = (ctypes.get(pid) or {}).get(wk)
        my_weeks = [w for w in weeks if ex["nights"][f"s{w['n']}"] > 0]
        txs = []
        for t in mov:
            if pid not in (t["from_id"], t["to_id"]):
                continue
            other = t["to_id"] if t["from_id"] == pid else t["from_id"]
            op = m["by_id"].get(other)
            if not op:
                err(f"Movimento n.{t['n']}: ID controparte {other} non in finance-map"); continue
            txs.append({"n": t["n"], "date": t["date"], "kind": t["kind"],
                        "direction": "out" if t["from_id"] == pid else "in",
                        "counterpart": op["public_name"], "amount": t["amount"],
                        "note": t["note"]})
        b = bon.get(pid) or {"quota_dovuta": 0.0, "quota_finale": 0.0, "versato": 0.0, "delta": 0.0}
        s = sal.get(pid) or {}
        if s and abs(s.get("quota_barca", 0) - b["quota_finale"]) > 0.01:
            warn(f"{p['public_name']}: quota barca saldi ({s.get('quota_barca')}) ≠ "
                 f"quota finale Bonifici ({b['quota_finale']}) — fa fede Bonifici")
        if s and abs(s.get("extra", 0) - ex["total"]) > 0.01:
            warn(f"{p['public_name']}: extra saldi ({s.get('extra')}) ≠ totale matrice Extra ({ex['total']})")
        payloads[cid] = {
            "generated_at": now,
            "name": p["public_name"],
            "dates": {"board": board, "leave": leave, "nights": ex["nights"]["tot"],
                      "weeks": [{"n": w["n"], "from": w["from"], "to": w["to"],
                                 "nights": ex["nights"][f"s{w['n']}"]} for w in my_weeks]},
            "cabins": ivals,
            "boat": {"quota_dovuta": b["quota_dovuta"], "quota_finale": b["quota_finale"]},
            "extras": ex["items"],
            "extras_total": ex["total"],
            "transactions": sorted(txs, key=lambda t: t["n"]),
            "totals": {"dovuto": s.get("dovuto", 0.0), "crediti": s.get("crediti", 0.0),
                       "versato": s.get("versato", 0.0), "ricevuto": s.get("ricevuto", 0.0),
                       "saldo": s.get("saldo", 0.0), "stato": s.get("stato") or None,
                       "delta_da_versare": b["delta"]},
        }
    return payloads, skipped


def upsert(payloads: dict[str, dict]) -> None:
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_KEY")
    if not url or not key:
        sys.exit("Servono le env-var SUPABASE_URL e SUPABASE_SERVICE_KEY (service role).")
    now = dt.datetime.now().astimezone().isoformat(timespec="seconds")
    rows = [{"crew_id": cid, "payload": pl, "updated_at": now}   # il default now() vale solo all'insert
            for cid, pl in sorted(payloads.items())]
    req = urllib.request.Request(
        url.rstrip("/") + "/rest/v1/voyage_finance",
        data=json.dumps(rows, ensure_ascii=False).encode(), method="POST",
        headers={"apikey": key, "Authorization": f"Bearer {key}",
                 "Content-Type": "application/json",
                 "Prefer": "resolution=merge-duplicates,return=minimal"})
    with urllib.request.urlopen(req, timeout=30) as r:
        print(f"voyage_finance: {len(rows)} righe upsert (HTTP {r.status})")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--json", action="store_true", help="col dry-run, stampa i payload completi")
    a = ap.parse_args()

    m = load_map()
    payloads, skipped = build_payloads(open_wb(), m)

    print(f"-- {len(payloads)} payload, {len(skipped)} esclusi --")
    for cid, pl in sorted(payloads.items()):
        t = pl["totals"]
        print(f"  {cid:<12} {pl['dates']['board']}→{pl['dates']['leave']} "
              f"{pl['dates']['nights']:>2}n  barca {pl['boat']['quota_finale']:>8.2f}  "
              f"extra {pl['extras_total']:>7.2f}  versato {t['versato']:>9.2f}  "
              f"delta {t['delta_da_versare']:>8.2f}  {t['stato'] or ''}")
    for s in skipped:
        print(f"  SKIPPED: {s}")
    for w in WARNINGS:
        print(f"  WARNING: {w}")
    for e in ERRORS:
        print(f"  ERRORE: {e}")
    if ERRORS:
        sys.exit(f"{len(ERRORS)} errori — nessun push.")
    if a.dry_run:
        if a.json:
            print(json.dumps(payloads, indent=2, ensure_ascii=False))
        print("-- dry run: niente push --")
    else:
        upsert(payloads)
