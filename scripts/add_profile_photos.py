# -*- coding: utf-8 -*-
"""Aggancia i RITRATTI dei passeggeri (avatar) da data/Profili/ ai profili del sito.

Metti in data/Profili/ un file per persona, nominato col NOME ("Leonardo Mascagni.jpg")
oppure col crew_id ("leonardo-m.jpg"). Lo strumento:
- risolve il crew_id (match diretto sull'id, oppure dal nome via registro Passengers);
- ottimizza il ritratto in site/crew/img/<id>.jpg (dedup: salta se e' gia' l'avatar attuale);
- lancia build_crew.py, che aggancia l'avatar (site/crew/img/<id>.jpg) ai profili alumni
  e aggiorna quelli 2026 (che puntano allo stesso file via crew2026.photo).

Uso:  python scripts/add_profile_photos.py            (applica)
      python scripts/add_profile_photos.py --dry      (anteprima)
"""
import openpyxl, collections, os, glob, re, unicodedata, json, sys, shutil, tempfile, subprocess
from PIL import Image, ImageOps

# ROOT = radice del repo (genitore di scripts/), relativo alla posizione dello
# script: cosi' funziona anche dai git worktree fuori OneDrive, non solo qui.
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DRY = "--dry" in sys.argv
CV = os.path.join(ROOT, "data", "Skipper CV Enriched.xlsx")
CREWJSON = os.path.join(ROOT, "site", "data", "crew.json")
PROFILIDIR = os.path.join(ROOT, "data", "Profili")
OUTDIR = os.path.join(ROOT, "site", "crew", "img")
IMG_EXT = (".jpg", ".jpeg", ".png", ".webp")
DEDUP_THR = 5

# --- registro passeggeri (per risolvere i nomi) — come build_trips.py ---
_cvtmp = os.path.join(tempfile.gettempdir(), "nina_cv_prof.xlsx")
try: shutil.copy(CV, _cvtmp)
except PermissionError:
    if not os.path.exists(_cvtmp): raise
wb = openpyxl.load_workbook(_cvtmp, data_only=True)
people = {}
for r in wb["Passengers"].iter_rows(min_row=2, values_only=True):
    if r[0] is None: continue
    people[int(r[0])] = {"cog": (r[1] or "").strip(), "nom": (r[2] or "").strip()}
part = collections.defaultdict(list)
for r in wb["Participations"].iter_rows(min_row=2, values_only=True):
    if r[0] is None or r[2] is None: continue
    try: t = int(r[0]); p = int(r[2])
    except: continue
    part[p].append(t)

def norm2(s): return "".join(c for c in unicodedata.normalize("NFD", (s or "").lower()) if c.isalnum())
def slug(s): return re.sub(r"[^a-z0-9]+", "-", unicodedata.normalize("NFKD", s.lower()).encode("ascii", "ignore").decode()).strip("-")
def norm(s): return "".join(c for c in unicodedata.normalize("NFD", s.lower()) if c.isalnum() or c == " ")
FIRSTFIX = {"jack": "giacomo", "bvernardo": "bernardo"}
ALIAS = {"edoardo-c": "edo-c", "gabriele-m": "gabri-m", "federico-b": "fede-b"}
MANUAL = {"edoardo casella": "edo-c"}

def resolve_pid(fullname):
    toks = fullname.split()
    if len(toks) < 2: return None
    first = FIRSTFIX.get(norm2(toks[0]), norm2(toks[0])); last = norm2(toks[-1])
    if len(first) < 2 or len(last) < 2: return None
    cands = [pid for pid, pp in people.items()
             if norm2(pp["nom"]).startswith(first[:4]) and norm2(pp["cog"]).startswith(last[:4])]
    if not cands: return None
    return max(cands, key=lambda p: len(part.get(p, [])))
def crew_id_of(pid):
    pp = people[pid]; cid = slug(pp["nom"] + " " + pp["cog"][:1])
    return ALIAS.get(cid, cid)

crew_ids = {p["id"] for p in json.load(open(CREWJSON, encoding="utf-8"))["people"]}

def resolve_avatar_id(stem):
    s = slug(stem)
    if s in crew_ids: return s                          # nominato direttamente col crew_id
    clean = re.sub(r"\([^)]*\)", " ", stem).replace("-", " ")
    clean = re.sub(r"(?i)\b(profile|profilo|avatar|ritratto|foto|img|dsc)\b", " ", clean)
    clean = re.sub(r"20[0-2][0-9]", " ", clean); clean = re.sub(r"\b\d+\b", " ", clean)
    clean = " ".join(clean.split())
    nk = " ".join(norm(clean).split())
    if nk in MANUAL: return MANUAL[nk]
    pid = resolve_pid(clean)
    return crew_id_of(pid) if pid else None

def save_avatar(im, path):
    mx = 800
    if max(im.size) > mx:
        sc = mx / max(im.size); im = im.resize((round(im.width * sc), round(im.height * sc)), Image.LANCZOS)
    im.save(path, "JPEG", quality=85, optimize=True, progressive=True)
    return im
def ahash(im):
    px = im.convert("L").resize((8, 8), Image.LANCZOS).tobytes(); avg = sum(px) / 64
    v = 0
    for i, p in enumerate(px):
        if p >= avg: v |= 1 << i
    return v
def ham(a, b): return bin(a ^ b).count("1")

applied, skipped, unresolved = [], [], []
for f in sorted(glob.glob(os.path.join(PROFILIDIR, "*"))):
    if not (os.path.isfile(f) and os.path.splitext(f)[1].lower() in IMG_EXT): continue
    stem = os.path.splitext(os.path.basename(f))[0]
    cid = resolve_avatar_id(stem)
    if not cid or cid not in crew_ids:
        unresolved.append((os.path.basename(f), cid or "nessun match")); continue
    try: im = ImageOps.exif_transpose(Image.open(f)).convert("RGB")
    except Exception as e:
        unresolved.append((os.path.basename(f), "illeggibile: %s" % e)); continue
    dst = os.path.join(OUTDIR, cid + ".jpg")
    if os.path.exists(dst):                              # gia' un avatar: applica solo se diverso
        try:
            if ham(ahash(im), ahash(ImageOps.exif_transpose(Image.open(dst)).convert("RGB"))) <= DEDUP_THR:
                skipped.append((cid, os.path.basename(f))); continue
        except Exception: pass
    if not DRY:
        os.makedirs(OUTDIR, exist_ok=True); save_avatar(im, dst)
    applied.append((cid, os.path.basename(f)))

print(("DRY-RUN — " if DRY else "") + "avatar applicati: %d, invariati: %d" % (len(applied), len(skipped)))
for cid, b in applied: print("  + %-14s <- %s" % (cid, b))
for cid, b in skipped: print("  = %-14s (gia' attuale) <- %s" % (cid, b))
for b, why in unresolved: print("  ! %s -> %s" % (b, why))
if not DRY and applied:
    print("\nAggancio gli avatar ai profili (build_crew.py)...")
    r = subprocess.run([sys.executable, os.path.join(ROOT, "scripts", "build_crew.py")], capture_output=True, text=True)
    print(r.stdout.strip() or r.stderr.strip())
