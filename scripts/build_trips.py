# Genera site/data/trips.json + ottimizza le foto viaggio in site/trips/img/<id>/.
# Legge le foto sia dalla root di data/Skipper Images/ sia dal sottofolder Crew/;
# le assegna ai viaggi per destinazione+anno, taggia le persone dal nome file
# (-> campo people nelle foto + sidecar data/photo-tags.json per build_crew.py) e
# ottimizza le 3 foto incidente per il dossier skipper.
#
# ATTENZIONE (2026-07-12): data/Skipper Images/ e' INCOMPLETA rispetto al sito (molte
# foto rinominate/rimosse) -> questo build ora produce ~15 foto invece delle 52 online.
# NON rieseguirlo: cancellerebbe ~37 foto + i loro tag dal sito. Per AGGIUNGERE foto
# nuove usa scripts/add_trip_photos.py (append-only, preserva le foto e i tag esistenti).
import openpyxl, collections, os, glob, re, unicodedata, json, sys

if "--force" not in sys.argv:
    sys.exit("BLOCCATO: data/Skipper Images/ e' incompleta rispetto al sito (vedi header):\n"
             "un rebuild cancellerebbe foto e tag pubblicati (successo il 2026-07-15, recuperato da git).\n"
             "Per aggiungere foto usa scripts/add_trip_photos.py; per rigenerare davvero tutto: --force.")
from PIL import Image, ImageOps

# ROOT = radice del repo (genitore di scripts/), relativo alla posizione dello
# script: cosi' funziona anche dai git worktree fuori OneDrive, non solo qui.
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CV = os.path.join(ROOT, "data", "Skipper CV Enriched.xlsx")
import shutil, tempfile
_cvtmp = os.path.join(tempfile.gettempdir(), "nina_cv_read.xlsx")
try:
    shutil.copy(CV, _cvtmp)  # copia: il file OneDrive/Excel puo' essere lockato
except PermissionError:
    if not os.path.exists(_cvtmp): raise   # nessuna copia utilizzabile -> chiudi Excel
    print("CV lockato da Excel/OneDrive: uso la copia temp esistente", _cvtmp)
CREWDIR = os.path.join(ROOT, "data", "Skipper Images", "Crew")
IMGROOT = os.path.join(ROOT, "data", "Skipper Images")
OUTIMG = os.path.join(ROOT, "site", "trips", "img")
STORIES = os.path.join(ROOT, "data", "trip-stories.json")   # racconti scritti a mano
stories = json.load(open(STORIES, encoding="utf-8")).get("stories", {}) if os.path.exists(STORIES) else {}
ZONE_OVERRIDE = {"cari": "Seychelles"}   # la zona "Inner Islands" della crociera Seychelles va mostrata come "Seychelles"

wb = openpyxl.load_workbook(_cvtmp, data_only=True)
people = {}
for r in wb["Passengers"].iter_rows(min_row=2, values_only=True):
    if r[0] is None: continue
    people[int(r[0])] = {"cog": (r[1] or "").strip(), "nom": (r[2] or "").strip()}
def ini(p):
    pp = people.get(p, {})
    return (pp.get("nom", "") + " " + (pp.get("cog", "")[:1] + "." if pp.get("cog") else "")).strip() or ("#" + str(p))

trips = {}
for r in wb["Summary"].iter_rows(min_row=2, values_only=True):
    if r[0] in (None, "TOTAL"): continue
    try: tid = int(float(r[0]))
    except: continue
    trips[tid] = {"year": int(float(r[1])), "month": r[2], "country": (r[4] or "").strip(),
                  "zone": (r[5] or "").strip(), "boat": (r[6] or "").strip(),
                  "name": (r[8] or "").strip(), "nm": int(float(r[9])), "days": float(r[10])}
crew = collections.defaultdict(list)   # trip id -> [passenger id]
for r in wb["Participations"].iter_rows(min_row=2, values_only=True):
    if r[0] is None or r[2] is None: continue
    try: t = int(r[0]); p = int(r[2])
    except: continue
    crew[t].append(p)
part = collections.defaultdict(list)   # passenger id -> [trip id]
for t, ps in crew.items():
    for p in ps: part[p].append(t)

# --- risoluzione nome taggato -> passenger id -> crew_id (id in crew.json) ---
def norm2(s): return "".join(c for c in unicodedata.normalize("NFD", (s or "").lower()) if c.isalnum())
def slug(s):
    return re.sub(r"[^a-z0-9]+", "-", unicodedata.normalize("NFKD", s.lower()).encode("ascii", "ignore").decode()).strip("-")
FIRSTFIX = {"jack": "giacomo", "bvernardo": "bernardo"}        # nickname / typo sul nome
ALIAS = {"edoardo-c": "edo-c", "gabriele-m": "gabri-m", "federico-b": "fede-b"}  # crew_id 2026 non derivabili
MANUAL = {"edoardo casella": "edo-c"}   # skipper: non e' nel registro Passengers, mappa diretta

def resolve_pid(fullname):
    toks = fullname.split()
    if len(toks) < 2: return None
    first = FIRSTFIX.get(norm2(toks[0]), norm2(toks[0])); last = norm2(toks[-1])
    if len(first) < 2 or len(last) < 2: return None
    cands = [pid for pid, pp in people.items()
             if norm2(pp["nom"]).startswith(first[:4]) and norm2(pp["cog"]).startswith(last[:4])]
    if not cands: return None
    return max(cands, key=lambda p: len(part.get(p, [])))   # tie-break: piu' partecipazioni
def crew_id_of(pid):
    pp = people[pid]; cid = slug(pp["nom"] + " " + pp["cog"][:1])
    return ALIAS.get(cid, cid)

# Overlay autoritativo per i viaggi con cambi di equipaggio. Il roster unificato
# resta disponibile per compatibilita', mentre crew_groups permette alla pagina
# viaggio di mostrare chiaramente ogni turno.
GROUPS_FILE = os.path.join(ROOT, "data", "trip-crew-groups.json")
group_cfg = json.load(open(GROUPS_FILE, encoding="utf-8")).get("trips", {}) if os.path.exists(GROUPS_FILE) else {}
display_by_cid = {crew_id_of(pid): ini(pid) for pid in people}
display_by_cid["edo-c"] = "Edo C"

KW = {"grecia": "Greece", "croazia": "Croazia", "spagna": "Spain", "sardegna": "Italy",
      "sicilia": "Italy", "grenadine": "Saint Vincent", "baleari": "Spain", "corsica": "France"}
YEAR_REMAP = {("grenadine", 2016): 2017}   # crociera di capodanno a cavallo d'anno
IMG_EXT = (".jpg", ".jpeg", ".png", ".webp")
STRIP = {"incidente", "incidenti", "naufragio", "di", "dei", "in", "porto",
         "trouble", "shooting", "pescatori", "napoletani", "agata", "leo", "e"}
DOSSIER = {"pescatori napoletani": "napoletani.jpg", "trouble shooting": "troubleshooting.jpg"}
# nota: la foto Moneglia e' gia' in site/skipper/img/moneglia.jpg (stessa sorgente) -> niente doppione
DEDUP_THR = 5   # distanza di Hamming aHash entro cui due foto dello stesso viaggio sono doppioni

def norm(s): return "".join(c for c in unicodedata.normalize("NFD", s.lower()) if c.isalnum() or c == " ")
def parse_people(basename):
    stem = os.path.splitext(basename)[0]
    pids, cids, unresolved = [], [], []
    for seg in re.split(r"\s+-\s+", stem):
        s = re.sub(r"20[0-2][0-9](-\d{1,2})?", " ", seg).replace("-", " ")
        for k in KW: s = re.sub(r"(?i)\b" + k + r"\b", " ", s)
        s = re.sub(r"(?i)\b(i{1,4}|iv|v)\b", " ", s)     # numeri romani (I..IIII usati)
        s = re.sub(r"\([^)]*\)", " ", s); s = re.sub(r"\b\d+\b", " ", s)
        s = " ".join(w for w in s.split() if w.lower() not in STRIP).strip()
        if len(s.split()) < 2: continue
        nk = " ".join(norm(s).split())
        if nk in MANUAL:                       # skipper (fuori registro)
            cids.append(MANUAL[nk]); continue
        pid = resolve_pid(s)
        if pid:
            pids.append(pid); cids.append(crew_id_of(pid))
        else:
            unresolved.append(s)
    return pids, cids, unresolved

# --- scan foto: root di Skipper Images + sottofolder Crew ---
srcs = [f for f in glob.glob(os.path.join(IMGROOT, "*")) + glob.glob(os.path.join(CREWDIR, "*"))
        if os.path.isfile(f) and os.path.splitext(f)[1].lower() in IMG_EXT]

photos = collections.defaultdict(list)   # trip id -> [{path, pids, cids, kw}]
skipped, unresolved_all = [], []
for f in sorted(srcs):
    b = os.path.basename(f); nl = norm(b); yr = re.search(r"(20[0-2][0-9])", b)
    kw = next((k for k in KW if k in nl), None)
    if not yr or not kw:
        if not b.lower().startswith(("dji_", "img_", "photo_", "whatsapp", "2016", "2017", "claudio-casella")):
            skipped.append((b, "no anno/destinazione"))
        continue
    if any(x in nl for x in ("moneglia", "incident", "trouble")) and "pescatori napoletani" not in nl:
        continue   # foto incidente (moneglia/trouble) fuori dalle gallerie viaggio; napoletani ammessa in sorbe
    y = int(yr.group(1)); yy = YEAR_REMAP.get((kw, y), y); country = KW[kw]
    cands = [t for t, m in trips.items() if m["year"] == yy and country in str(m["country"])]
    if country == "Italy":
        zk = "sardegna" if "sardegna" in nl else ("sicil" if "sicilia" in nl else None)
        cands = [t for t in cands if zk and zk in norm(str(trips[t]["zone"]))]
    pids, cids, unres = parse_people(b)
    if unres: unresolved_all.append((b, unres))
    tgt = None
    if len(cands) == 1:
        tgt = cands[0]
    elif len(cands) > 1 and pids:   # ambiguo (stesso anno/zona) -> roster ciurma, poi prevalenza del cluster
        sc = sorted(cands, key=lambda t: (sum(p in crew[t] for p in pids), len(photos[t])), reverse=True)
        if sum(p in crew[sc[0]] for p in pids) > 0: tgt = sc[0]
    if tgt is None:
        skipped.append((b, "nessun viaggio" if not cands else "ambiguo (%d candidati)" % len(cands)))
        continue
    photos[tgt].append({"path": f, "cids": cids, "kw": kw})

# --- ottimizzazione immagini + dedup per viaggio ---
def save_opt(im, path):
    mx = 1280
    if max(im.size) > mx:
        sc = mx / max(im.size); im = im.resize((round(im.width * sc), round(im.height * sc)), Image.LANCZOS)
    im.save(path, "JPEG", quality=82, optimize=True, progressive=True)
    return im
def ahash(im):
    px = im.convert("L").resize((8, 8), Image.LANCZOS).tobytes(); avg = sum(px) / 64
    v = 0
    for i, p in enumerate(px):
        if p >= avg: v |= 1 << i
    return v
def ham(a, b): return bin(a ^ b).count("1")

out, dedup_log = [], []
for t in sorted(trips):
    m = trips[t]; tid = slug(m["name"]) or ("t" + str(t))
    group_spec = group_cfg.get(tid)
    crew_groups_out = []
    if group_spec:
        if group_spec.get("trip_id") is not None and int(group_spec["trip_id"]) != t:
            raise ValueError(f"Configurazione equipaggi {tid}: trip_id {group_spec['trip_id']} non corrisponde a {t}")
        for g in group_spec.get("groups", []):
            unknown = [cid for cid in g.get("crew", []) if cid not in display_by_cid]
            if unknown:
                raise ValueError(f"{tid}/{g.get('id')}: crew id sconosciuti: {', '.join(unknown)}")
            crew_groups_out.append({"id": g["id"], "label": g["label"], "days": g["days"],
                                    "crew": [display_by_cid[cid] for cid in g["crew"]]})
        configured = list(dict.fromkeys(cid for g in group_spec.get("groups", []) for cid in g.get("crew", []) if cid != "edo-c"))
        trip_crew = [display_by_cid[cid] for cid in configured]
    else:
        trip_crew = [ini(p) for p in crew.get(t, [])]
    phs = []
    items = photos.get(t, [])
    if items:
        d = os.path.join(OUTIMG, tid)
        if os.path.isdir(d):
            for old in glob.glob(os.path.join(d, "*")): os.remove(old)   # niente orfani da run precedenti
        os.makedirs(d, exist_ok=True)
        kept = []; n = 0   # kept: [(ahash, indice in phs)]
        for it in sorted(items, key=lambda x: x["path"]):
            im = ImageOps.exif_transpose(Image.open(it["path"])).convert("RGB")
            h = ahash(im)
            di = next((ki for hh, ki in kept if ham(hh, h) <= DEDUP_THR), None)
            if di is not None:   # doppione: unisci i tag alla copia tenuta, non riscrivere
                for c in it["cids"]:
                    if c not in phs[di]["people"]: phs[di]["people"].append(c)
                dedup_log.append((tid, os.path.basename(it["path"])))
                continue
            n += 1; fn = "%02d.jpg" % n
            im = save_opt(im, os.path.join(d, fn))
            phs.append({"src": "trips/img/%s/%s" % (tid, fn), "w": im.width, "h": im.height,
                        "people": list(dict.fromkeys(it["cids"])), "caption": it["kw"].title() + " " + str(m["year"])})
            kept.append((h, len(phs) - 1))
    out.append({"id": tid, "year": m["year"], "month": m["month"], "country": m["country"],
                "zone": ZONE_OVERRIDE.get(tid, m["zone"]), "boat": m["boat"], "name": m["name"], "nm": m["nm"],
                "days": round(m["days"], 1), "crew": trip_crew,
                **({"crew_groups": crew_groups_out} if crew_groups_out else {}),
                "n_photos": len(phs), "photos": phs,
                **({"story": stories[tid]} if tid in stories else {})})

# --- foto incidente per il dossier skipper (site/skipper/img/<slug>.jpg) ---
SKIMG = os.path.join(ROOT, "site", "skipper", "img"); os.makedirs(SKIMG, exist_ok=True)
dossier_done = []
for f in sorted(srcs):
    nl = norm(os.path.basename(f))
    for sub, outname in DOSSIER.items():
        if sub in nl:
            save_opt(ImageOps.exif_transpose(Image.open(f)).convert("RGB"), os.path.join(SKIMG, outname))
            dossier_done.append(outname); break

# ripartizione miglia per paese sui viaggi multi-paese, da data/miles-by-country.json
# (generata da scripts/miles_by_country.py dai tracciati GPS Google Earth). Additivo,
# per id: preserva nm_by_country anche se questo build viene rieseguito.
MBC = os.path.join(ROOT, "data", "miles-by-country.json")
if os.path.exists(MBC):
    mbc = json.load(open(MBC, encoding="utf-8")).get("trips", {})
    for t in out:
        if t["id"] in mbc:
            t["nm_by_country"] = mbc[t["id"]]["nm_by_country"]

data = {"generated_at": "2026-07-11", "trips": out}
with open(os.path.join(ROOT, "site", "data", "trips.json"), "w", encoding="utf-8") as fh:
    json.dump(data, fh, ensure_ascii=False, indent=2)

# --- sidecar: foto per persona (letto da build_crew.py) ---
by_person = collections.defaultdict(list)
for t in out:
    for p in t["photos"]:
        for cid in p["people"]:
            by_person[cid].append({"src": p["src"], "trip_id": t["id"], "caption": p["caption"]})
with open(os.path.join(ROOT, "data", "photo-tags.json"), "w", encoding="utf-8") as fh:
    json.dump({"generated_at": "2026-07-11", "by_person": by_person}, fh, ensure_ascii=False, indent=1)

# --- log ---
tot = sum(t["n_photos"] for t in out)
print("trips.json:", len(out), "viaggi,", tot, "foto ottimizzate,", len(by_person), "persone taggate,",
      sum(1 for t in out if t.get("story")), "racconti")
miss_story = [t["id"] for t in out if not t.get("story")]
if miss_story: print("viaggi senza racconto:", ", ".join(miss_story))
for t in out:
    if t["n_photos"]:
        ppl = sorted({c for p in t["photos"] for c in p["people"]})
        print("  %-12s %s %-16s -> %2d foto  tag: %s" % (t["id"], t["year"], t["country"], t["n_photos"], ", ".join(ppl) or "-"))
if dossier_done: print("dossier skipper:", ", ".join(dossier_done))
if dedup_log: print("doppioni scartati (dedup):", "; ".join("%s/%s" % x for x in dedup_log))
if unresolved_all:
    print("nomi NON risolti (foto comunque assegnata al viaggio):")
    for b, u in unresolved_all: print("  ", b, "->", u)
if skipped:
    print("foto NON assegnate:")
    for b, why in skipped: print("  ", b, "->", why)
